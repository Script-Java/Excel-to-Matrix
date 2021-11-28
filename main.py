import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook


st.write("""
## Excel To Matrix Converter

Please make sure to have your excel file in the same folder as the script
""")


def file_selector(folder_path='.'):
    filenames = os.listdir(folder_path)
    selected_filename = st.selectbox('Select a file', filenames)
    return os.path.join(folder_path, selected_filename)

filename = file_selector()
st.write('You selected `%s`' % filename)

try:
    data = pd.read_excel(filename)
    st.write(data)
except:
    st.write("File Type Not Supported")
    
matrix = [[1,2,3],[4,5,6],[7,8,9]]
reactance_values = []

def generate_matrix():
    wb = load_workbook(filename)
    sheet = wb.sheetnames
    ws = wb[sheet[1]]
    reactance = ws["D"]
    for cell in reactance:
        if type(cell.value) == float:
            reactance_values.append(cell.value)
    matrix[0][1] = 1/reactance_values[0]
    matrix[0][2] = 1/reactance_values[1]
    matrix[0][0] = matrix[0][1] + matrix[0][2]
    # First Row Above
    matrix[1][0] = 1/reactance_values[0]
    matrix[1][2] = 1/reactance_values[2]
    matrix[1][1] = matrix[1][0] + matrix[1][2]
    # Second Row Above
    matrix[2][0] = 1/reactance_values[1]
    matrix[2][1] = 1/reactance_values[2]
    matrix[2][2] = matrix[2][0] + matrix[2][1]
    # Third Row Above
    return matrix
new_matrix = generate_matrix()
st.write(new_matrix)